"""
Pullus QA Tracker - Excel Template Generator
Creates a beautiful, professional Excel template for poultry processing QA tracking
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

def create_qa_tracker_template():
    """Generate the complete QA Tracker Excel template"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create all three sheets
    create_pullus_qa_tracker_sheet(wb)
    create_blast_freezing_tracker_sheet(wb)
    create_definition_of_terms_sheet(wb)

    # Save the workbook
    filename = 'Pullus_QA_Tracker_Template.xlsx'
    wb.save(filename)
    print(f"✓ Template created successfully: {filename}")
    return filename


def create_pullus_qa_tracker_sheet(wb):
    """Create the main Pullus QA Tracker sheet"""

    ws = wb.create_sheet("Pullus QA Tracker", 0)

    # Define column headers
    headers = [
        "Date",
        "Purchase Officer",
        "Location",
        "Number of Birds",
        "Number of Slaughter Men",
        "Slaughter Start Time",
        "Slaughter End Time",
        "Slaughter Duration",
        "Slaughter Status",
        "Weighing Start Time",
        "Weighing Stop Time",
        "Weighing Duration",
        "Weighing Status",
        "Invoice Writing Start Time",
        "Invoice Writing Stop Time",
        "Invoice Writing Duration",
        "Invoice Writing Status",
        "Transportation Mode",
        "Logistics Pickup Time",
        "Receipt to Pickup Duration",
        "Logistics Status",
        "Cold Room Arrival Time",
        "Pickup to Cold Room Duration",
        "Transportation Status",
        "Number of Washing Personnel",
        "Washing Start Time",
        "Washing End Time",
        "Washing Duration",
        "Washing Status",
        "Blast Freezer Time",
        "Total Process Duration",
        "Comments"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set column widths
    column_widths = {
        'A': 12,  # Date
        'B': 18,  # Purchase Officer
        'C': 15,  # Location
        'D': 14,  # Number of Birds
        'E': 14,  # Number of Slaughter Men
        'F': 16,  # Slaughter Start Time
        'G': 16,  # Slaughter End Time
        'H': 16,  # Slaughter Duration
        'I': 16,  # Slaughter Status
        'J': 16,  # Weighing Start Time
        'K': 16,  # Weighing Stop Time
        'L': 15,  # Weighing Duration
        'M': 15,  # Weighing Status
        'N': 18,  # Invoice Writing Start Time
        'O': 18,  # Invoice Writing Stop Time
        'P': 18,  # Invoice Writing Duration
        'Q': 18,  # Invoice Writing Status
        'R': 18,  # Transportation Mode
        'S': 16,  # Logistics Pickup Time
        'T': 18,  # Receipt to Pickup Duration
        'U': 15,  # Logistics Status
        'V': 18,  # Cold Room Arrival Time
        'W': 20,  # Pickup to Cold Room Duration
        'X': 18,  # Transportation Status
        'Y': 18,  # Number of Washing Personnel
        'Z': 16,  # Washing Start Time
        'AA': 16, # Washing End Time
        'AB': 15, # Washing Duration
        'AC': 15, # Washing Status
        'AD': 16, # Blast Freezer Time
        'AE': 18, # Total Process Duration
        'AF': 20  # Comments
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Add dropdown validations
    add_pullus_qa_dropdowns(ws)

    # Add demo data (rows 2-6)
    add_demo_data_pullus_qa(ws)

    # Add formulas starting from row 2 for 2000 rows
    add_pullus_qa_formulas(ws, start_row=2, num_rows=2000)

    # Apply styling to data rows (2000 rows)
    style_pullus_qa_data_rows(ws, start_row=2, num_rows=2000)

    return ws


def add_pullus_qa_dropdowns(ws):
    """Add dropdown validations to Pullus QA Tracker"""

    # Purchase Officer dropdown (Column B)
    purchase_officers = [
        "Abdulrasheed Rufai",
        "Theophilus Bulus",
        "Femi Abubakar",
        "Simon Agbosua",
        "Friday Shehu",
        "Victor Adeojo",
        "Alexander Baton",
        "Sabastine Peter"
    ]
    dv_purchase_officer = DataValidation(
        type="list",
        formula1=f'"{",".join(purchase_officers)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_purchase_officer)
    dv_purchase_officer.add(f'B2:B2001')

    # Transportation Mode dropdown (Column R)
    transport_modes = ["Keke", "Van", "Commercial Bus", "Pullus Bus", "Bike"]
    dv_transport = DataValidation(
        type="list",
        formula1=f'"{",".join(transport_modes)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_transport)
    dv_transport.add(f'R2:R2001')

    # Logistics Status dropdown (Column U)
    logistics_status = ["Ontime", "Early", "Delayed", "Very Delayed"]
    dv_logistics = DataValidation(
        type="list",
        formula1=f'"{",".join(logistics_status)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_logistics)
    dv_logistics.add(f'U2:U2001')

    # Transportation Status dropdown (Column X)
    transport_status = ["Good", "Manageable", "Bad", "Dangerous"]
    dv_transport_status = DataValidation(
        type="list",
        formula1=f'"{",".join(transport_status)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_transport_status)
    dv_transport_status.add(f'X2:X2001')


def add_demo_data_pullus_qa(ws):
    """Add realistic demo data to Pullus QA Tracker"""

    demo_data = [
        {
            'date': datetime(2025, 1, 15),
            'officer': 'Abdulrasheed Rufai',
            'location': 'Ikeja Processing Center',
            'birds': 500,
            'slaughter_men': 4,
            'slaughter_start': '6:30 AM',
            'slaughter_end': '8:15 AM',  # 1h 45m - Good
            'weighing_start': '8:20 AM',
            'weighing_stop': '8:24 AM',  # 4 min - Good
            'invoice_start': '8:25 AM',
            'invoice_stop': '8:29 AM',  # 4 min - Good
            'transport_mode': 'Pullus Bus',
            'pickup_time': '8:35 AM',
            'logistics_status': 'Ontime',
            'coldroom_arrival': '9:15 AM',
            'transport_status': 'Good',
            'washing_personnel': 3,
            'washing_start': '9:20 AM',
            'washing_end': '9:50 AM',  # 30 min - Good
            'blast_time': '10:00 AM',
            'comments': 'Smooth operation, all targets met'
        },
        {
            'date': datetime(2025, 1, 16),
            'officer': 'Theophilus Bulus',
            'location': 'Surulere Hub',
            'birds': 750,
            'slaughter_men': 3,
            'slaughter_start': '7:00 AM',
            'slaughter_end': '9:30 AM',  # 2h 30m - Manageable
            'weighing_start': '9:40 AM',
            'weighing_stop': '9:48 AM',  # 8 min - Manageable
            'invoice_start': '9:50 AM',
            'invoice_stop': '9:58 AM',  # 8 min - Manageable
            'transport_mode': 'Van',
            'pickup_time': '10:05 AM',
            'logistics_status': 'Delayed',
            'coldroom_arrival': '10:45 AM',
            'transport_status': 'Manageable',
            'washing_personnel': 2,
            'washing_start': '10:50 AM',
            'washing_end': '11:30 AM',  # 40 min - Manageable
            'blast_time': '11:45 AM',
            'comments': 'Slight delay in logistics pickup'
        },
        {
            'date': datetime(2025, 1, 17),
            'officer': 'Femi Abubakar',
            'location': 'Yaba Station',
            'birds': 300,
            'slaughter_men': 2,
            'slaughter_start': '8:00 AM',
            'slaughter_end': '9:45 AM',  # 1h 45m - Good
            'weighing_start': '9:50 AM',
            'weighing_stop': '9:53 AM',  # 3 min - Good
            'invoice_start': '9:54 AM',
            'invoice_stop': '9:57 AM',  # 3 min - Good
            'transport_mode': 'Keke',
            'pickup_time': '10:00 AM',
            'logistics_status': 'Ontime',
            'coldroom_arrival': '10:25 AM',
            'transport_status': 'Good',
            'washing_personnel': 2,
            'washing_start': '10:30 AM',
            'washing_end': '10:55 AM',  # 25 min - Good
            'blast_time': '11:10 AM',
            'comments': 'Excellent timing throughout'
        },
        {
            'date': datetime(2025, 1, 18),
            'officer': 'Simon Agbosua',
            'location': 'Agege Processing Hub',
            'birds': 1000,
            'slaughter_men': 5,
            'slaughter_start': '5:30 AM',
            'slaughter_end': '9:00 AM',  # 3h 30m - Bad
            'weighing_start': '9:15 AM',
            'weighing_stop': '9:28 AM',  # 13 min - Bad
            'invoice_start': '9:30 AM',
            'invoice_stop': '9:43 AM',  # 13 min - Bad
            'transport_mode': 'Commercial Bus',
            'pickup_time': '9:50 AM',
            'logistics_status': 'Very Delayed',
            'coldroom_arrival': '11:00 AM',
            'transport_status': 'Bad',
            'washing_personnel': 4,
            'washing_start': '11:10 AM',
            'washing_end': '12:05 PM',  # 55 min - Bad
            'blast_time': '12:20 PM',
            'comments': 'High volume caused delays, need more personnel'
        },
        {
            'date': datetime(2025, 1, 19),
            'officer': 'Friday Shehu',
            'location': 'Oshodi Center',
            'birds': 600,
            'slaughter_men': 3,
            'slaughter_start': '6:45 AM',
            'slaughter_end': '8:30 AM',  # 1h 45m - Good
            'weighing_start': '8:35 AM',
            'weighing_stop': '8:41 AM',  # 6 min - Manageable
            'invoice_start': '8:42 AM',
            'invoice_stop': '8:48 AM',  # 6 min - Manageable
            'transport_mode': 'Pullus Bus',
            'pickup_time': '8:50 AM',
            'logistics_status': 'Early',
            'coldroom_arrival': '9:20 AM',
            'transport_status': 'Good',
            'washing_personnel': 3,
            'washing_start': '9:25 AM',
            'washing_end': '9:58 AM',  # 33 min - Manageable
            'blast_time': '10:10 AM',
            'comments': 'Good performance overall'
        }
    ]

    for row_idx, data in enumerate(demo_data, start=2):
        ws[f'A{row_idx}'] = data['date']
        ws[f'A{row_idx}'].number_format = 'DD-MMM-YYYY'

        ws[f'B{row_idx}'] = data['officer']
        ws[f'C{row_idx}'] = data['location']
        ws[f'D{row_idx}'] = data['birds']
        ws[f'E{row_idx}'] = data['slaughter_men']

        # Time entries
        ws[f'F{row_idx}'] = data['slaughter_start']
        ws[f'G{row_idx}'] = data['slaughter_end']
        ws[f'J{row_idx}'] = data['weighing_start']
        ws[f'K{row_idx}'] = data['weighing_stop']
        ws[f'N{row_idx}'] = data['invoice_start']
        ws[f'O{row_idx}'] = data['invoice_stop']

        ws[f'R{row_idx}'] = data['transport_mode']
        ws[f'S{row_idx}'] = data['pickup_time']
        ws[f'U{row_idx}'] = data['logistics_status']
        ws[f'V{row_idx}'] = data['coldroom_arrival']
        ws[f'X{row_idx}'] = data['transport_status']

        ws[f'Y{row_idx}'] = data['washing_personnel']
        ws[f'Z{row_idx}'] = data['washing_start']
        ws[f'AA{row_idx}'] = data['washing_end']
        ws[f'AD{row_idx}'] = data['blast_time']

        ws[f'AF{row_idx}'] = data['comments']


def add_pullus_qa_formulas(ws, start_row, num_rows):
    """Add calculation formulas to Pullus QA Tracker"""

    for row in range(start_row, start_row + num_rows):
        # H: Slaughter Duration (G - F) - Format: "1hr 23min"
        ws[f'H{row}'] = (
            f'=IF(AND(G{row}<>"",F{row}<>""),'
            f'TEXT(INT((G{row}-F{row})*24),"0") & "hr " & TEXT(MOD((G{row}-F{row})*24*60,60),"0") & "min","")'
        )

        # I: Slaughter Status (based on H duration in hours)
        ws[f'I{row}'] = (
            f'=IF(AND(G{row}<>"",F{row}<>""),'
            f'IF((G{row}-F{row})<=TIME(2,0,0),"Good",'
            f'IF((G{row}-F{row})<=TIME(3,0,0),"Manageable",'
            f'IF((G{row}-F{row})<=TIME(4,0,0),"Bad","Dangerous"))),"")'
        )

        # L: Weighing Duration (K - J) - Format: "0hr 5min"
        ws[f'L{row}'] = (
            f'=IF(AND(K{row}<>"",J{row}<>""),'
            f'TEXT(INT((K{row}-J{row})*24),"0") & "hr " & TEXT(MOD((K{row}-J{row})*24*60,60),"0") & "min","")'
        )

        # M: Weighing Status (based on L duration in minutes)
        ws[f'M{row}'] = (
            f'=IF(AND(K{row}<>"",J{row}<>""),'
            f'IF((K{row}-J{row})<=TIME(0,5,0),"Good",'
            f'IF((K{row}-J{row})<=TIME(0,10,0),"Manageable",'
            f'IF((K{row}-J{row})<=TIME(0,15,0),"Bad","Dangerous"))),"")'
        )

        # P: Invoice Writing Duration (O - N) - Format: "0hr 5min"
        ws[f'P{row}'] = (
            f'=IF(AND(O{row}<>"",N{row}<>""),'
            f'TEXT(INT((O{row}-N{row})*24),"0") & "hr " & TEXT(MOD((O{row}-N{row})*24*60,60),"0") & "min","")'
        )

        # Q: Invoice Writing Status (based on P duration in minutes)
        ws[f'Q{row}'] = (
            f'=IF(AND(O{row}<>"",N{row}<>""),'
            f'IF((O{row}-N{row})<=TIME(0,5,0),"Good",'
            f'IF((O{row}-N{row})<=TIME(0,10,0),"Manageable",'
            f'IF((O{row}-N{row})<=TIME(0,15,0),"Bad","Dangerous"))),"")'
        )

        # T: Receipt to Pickup Duration (S - O) - Format: "1hr 23min"
        ws[f'T{row}'] = (
            f'=IF(AND(S{row}<>"",O{row}<>""),'
            f'TEXT(INT((S{row}-O{row})*24),"0") & "hr " & TEXT(MOD((S{row}-O{row})*24*60,60),"0") & "min","")'
        )

        # W: Pickup to Cold Room Duration (V - S) - Format: "1hr 23min"
        ws[f'W{row}'] = (
            f'=IF(AND(V{row}<>"",S{row}<>""),'
            f'TEXT(INT((V{row}-S{row})*24),"0") & "hr " & TEXT(MOD((V{row}-S{row})*24*60,60),"0") & "min","")'
        )

        # AB: Washing Duration (AA - Z) - Format: "0hr 47min"
        ws[f'AB{row}'] = (
            f'=IF(AND(AA{row}<>"",Z{row}<>""),'
            f'TEXT(INT((AA{row}-Z{row})*24),"0") & "hr " & TEXT(MOD((AA{row}-Z{row})*24*60,60),"0") & "min","")'
        )

        # AC: Washing Status (based on AB duration in minutes)
        ws[f'AC{row}'] = (
            f'=IF(AND(AA{row}<>"",Z{row}<>""),'
            f'IF((AA{row}-Z{row})<=TIME(0,30,0),"Good",'
            f'IF((AA{row}-Z{row})<=TIME(0,45,0),"Manageable","Bad")),"")'
        )

        # AE: Total Process Duration (AD - F) - Format: "3hr 24min"
        ws[f'AE{row}'] = (
            f'=IF(AND(AD{row}<>"",F{row}<>""),'
            f'TEXT(INT((AD{row}-F{row})*24),"0") & "hr " & TEXT(MOD((AD{row}-F{row})*24*60,60),"0") & "min","")'
        )


def style_pullus_qa_data_rows(ws, start_row, num_rows):
    """Apply styling to data rows in Pullus QA Tracker"""

    # Alternating row colors
    light_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    for row in range(start_row, start_row + num_rows):
        fill = light_fill if row % 2 == 0 else white_fill

        for col in range(1, 33):  # 32 columns
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Center align numeric columns
    numeric_cols = ['D', 'E', 'Y']
    for col in numeric_cols:
        for row in range(start_row, start_row + num_rows):
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")

    # Color-code status columns
    status_columns = ['I', 'M', 'Q', 'AC']  # Slaughter Status, Weighing Status, Invoice Writing Status, Washing Status
    for col in status_columns:
        for row in range(start_row, start_row + num_rows):
            cell = ws[f'{col}{row}']
            # Note: Conditional formatting would be ideal but requires more complex openpyxl setup
            # For now, we'll rely on the formulas
            cell.alignment = Alignment(horizontal="center", vertical="center")


def create_blast_freezing_tracker_sheet(wb):
    """Create the Blast Freezing Tracker sheet"""

    ws = wb.create_sheet("Blast Freezing Tracker", 1)

    # Define column headers
    headers = [
        "Date",
        "Arrival Time",
        "Blast Start Time",
        "Blast End Time",
        "Blast Duration",
        "Qty",
        "Target Temperature (°C)",
        "Actual Temperature (°C)",
        "Temperature Status"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set column widths
    column_widths = {'A': 12, 'B': 14, 'C': 16, 'D': 16, 'E': 14, 'F': 10, 'G': 20, 'H': 20, 'I': 16}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add demo data
    add_demo_data_blast_freezing(ws)

    # Add formulas for 2000 rows
    add_blast_freezing_formulas(ws, start_row=2, num_rows=2000)

    # Style data rows (2000 rows)
    style_blast_freezing_data_rows(ws, start_row=2, num_rows=2000)

    return ws


def add_demo_data_blast_freezing(ws):
    """Add demo data to Blast Freezing Tracker"""

    demo_data = [
        {
            'date': datetime(2025, 1, 15),
            'arrival': '9:15 AM',
            'blast_start': '9:30 AM',
            'blast_end': '2:30 PM',
            'qty': 500,
            'target_temp': -20,
            'actual_temp': -19
        },
        {
            'date': datetime(2025, 1, 16),
            'arrival': '10:45 AM',
            'blast_start': '11:00 AM',
            'blast_end': '4:30 PM',
            'qty': 750,
            'target_temp': -20,
            'actual_temp': -21
        },
        {
            'date': datetime(2025, 1, 17),
            'arrival': '10:25 AM',
            'blast_start': '10:40 AM',
            'blast_end': '3:40 PM',
            'qty': 300,
            'target_temp': -20,
            'actual_temp': -20
        }
    ]

    for row_idx, data in enumerate(demo_data, start=2):
        ws[f'A{row_idx}'] = data['date']
        ws[f'A{row_idx}'].number_format = 'DD-MMM-YYYY'

        ws[f'B{row_idx}'] = data['arrival']
        ws[f'C{row_idx}'] = data['blast_start']
        ws[f'D{row_idx}'] = data['blast_end']
        ws[f'F{row_idx}'] = data['qty']
        ws[f'G{row_idx}'] = data['target_temp']
        ws[f'H{row_idx}'] = data['actual_temp']


def add_blast_freezing_formulas(ws, start_row, num_rows):
    """Add formulas to Blast Freezing Tracker"""

    for row in range(start_row, start_row + num_rows):
        # Set default target temperature
        if ws[f'G{row}'].value is None:
            ws[f'G{row}'] = -20

        # E: Blast Duration (D - C) - Format: "5hr 0min"
        ws[f'E{row}'] = (
            f'=IF(AND(D{row}<>"",C{row}<>""),'
            f'TEXT(INT((D{row}-C{row})*24),"0") & "hr " & TEXT(MOD((D{row}-C{row})*24*60,60),"0") & "min","")'
        )

        # I: Temperature Status
        ws[f'I{row}'] = (
            f'=IF(AND(H{row}<>"",G{row}<>""),'
            f'IF(AND(H{row}>=-22,H{row}<=-18),"Good","Bad"),"")'
        )


def style_blast_freezing_data_rows(ws, start_row, num_rows):
    """Apply styling to Blast Freezing Tracker data rows"""

    light_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    for row in range(start_row, start_row + num_rows):
        fill = light_fill if row % 2 == 0 else white_fill

        for col in range(1, 10):  # 9 columns
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Temperature formatting
    for row in range(start_row, start_row + num_rows):
        ws[f'G{row}'].number_format = '0"°C"'
        ws[f'H{row}'].number_format = '0"°C"'


def create_definition_of_terms_sheet(wb):
    """Create the Definition of Terms documentation sheet"""

    ws = wb.create_sheet("Definition of Terms", 2)

    # Headers
    headers = ["Column Name", "Description", "Type / Values"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="6B4C9A", end_color="6B4C9A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # Freeze header
    ws.freeze_panes = 'A2'

    # Add definitions for Pullus QA Tracker
    definitions = [
        ("", "PULLUS QA TRACKER", ""),
        ("Date", "Date of processing operation", "Date format: DD-MMM-YYYY"),
        ("Purchase Officer", "Officer responsible for the purchase operation", "Dropdown: 8 officers"),
        ("Location", "Processing location/center", "Text entry"),
        ("Number of Birds", "Quantity of birds processed", "Number"),
        ("Number of Slaughter Men", "Personnel count for slaughter", "Number"),
        ("Slaughter Start Time", "When slaughter operation begins", "Time (12-hour AM/PM)"),
        ("Slaughter End Time", "When slaughter operation ends", "Time (12-hour AM/PM)"),
        ("Slaughter Duration", "Time taken for slaughter", "Auto-calculated format: 1hr 23min"),
        ("Slaughter Status", "Quality status of slaughter timing", "Auto: Good(≤2h), Manageable(2-3h), Bad(3-4h), Dangerous(>4h)"),
        ("Weighing Start Time", "When bird weighing begins", "Time (12-hour AM/PM)"),
        ("Weighing Stop Time", "When bird weighing ends", "Time (12-hour AM/PM)"),
        ("Weighing Duration", "Time taken for weighing", "Auto-calculated format: 0hr 5min"),
        ("Weighing Status", "Quality status of weighing timing", "Auto: Good(≤5min), Manageable(5-10min), Bad(10-15min), Dangerous(>15min)"),
        ("Invoice Writing Start Time", "When invoice writing begins", "Time (12-hour AM/PM)"),
        ("Invoice Writing Stop Time", "When invoice writing ends", "Time (12-hour AM/PM)"),
        ("Invoice Writing Duration", "Time taken for invoice writing", "Auto-calculated format: 0hr 5min"),
        ("Invoice Writing Status", "Quality status of invoice writing timing", "Auto: Good(≤5min), Manageable(5-10min), Bad(10-15min), Dangerous(>15min)"),
        ("Transportation Mode", "Vehicle type used for transport", "Dropdown: Keke, Van, Commercial Bus, Pullus Bus, Bike"),
        ("Logistics Pickup Time", "When logistics picks up birds", "Time (12-hour AM/PM)"),
        ("Receipt to Pickup Duration", "Time from invoice completion to logistics pickup", "Auto-calculated format: 0hr 47min"),
        ("Logistics Status", "Timeliness of logistics pickup", "Manual Dropdown: Ontime, Early, Delayed, Very Delayed"),
        ("Cold Room Arrival Time", "When birds arrive at cold storage", "Time (12-hour AM/PM)"),
        ("Pickup to Cold Room Duration", "Transport time to cold room", "Auto-calculated format: 1hr 9min"),
        ("Transportation Status", "Quality of transportation timing", "Manual Dropdown: Good, Manageable, Bad, Dangerous"),
        ("Number of Washing Personnel", "Personnel count for washing", "Number"),
        ("Washing Start Time", "When washing begins", "Time (12-hour AM/PM)"),
        ("Washing End Time", "When washing ends", "Time (12-hour AM/PM)"),
        ("Washing Duration", "Time taken for washing", "Auto-calculated format: 0hr 30min"),
        ("Washing Status", "Quality status of washing timing", "Auto: Good(≤30min), Manageable(31-45min), Bad(>45min)"),
        ("Blast Freezer Time", "When birds enter blast freezer", "Time (12-hour AM/PM)"),
        ("Total Process Duration", "Total time from start to blast freezer", "Auto-calculated format: 3hr 24min"),
        ("Comments", "Additional notes and observations", "Text entry"),
        ("", "", ""),
        ("", "BLAST FREEZING TRACKER", ""),
        ("Date", "Date of blast freezing", "Date format: DD-MMM-YYYY"),
        ("Arrival Time", "When birds arrive at cold room", "Time (12-hour AM/PM)"),
        ("Blast Start Time", "When blast freezing begins", "Time (12-hour AM/PM)"),
        ("Blast End Time", "When blast freezing ends", "Time (12-hour AM/PM)"),
        ("Blast Duration", "Time in blast freezer", "Auto-calculated format: 5hr 0min"),
        ("Qty", "Quantity of birds in batch", "Number"),
        ("Target Temperature (°C)", "Target blast temperature", "Number (default -20°C)"),
        ("Actual Temperature (°C)", "Actual recorded temperature", "Number with °C"),
        ("Temperature Status", "Temperature quality check", "Auto: Good(-22°C to -18°C), Bad(outside range)")
    ]

    row_num = 2
    for definition in definitions:
        if definition[0] == "":  # Section header
            cell = ws.cell(row=row_num, column=1)
            cell.value = definition[1]
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f'A{row_num}:C{row_num}')
        else:
            for col_num, value in enumerate(definition, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        row_num += 1

    # Set row heights for better readability
    for row in range(2, row_num):
        ws.row_dimensions[row].height = 25

    return ws


if __name__ == "__main__":
    create_qa_tracker_template()
